import openpyxl
import json

wb = openpyxl.load_workbook(
    'data/K014749_APQC Process Classification Framework (PCF) - Cross-Industry - Excel Version 7.4.xlsx',
    read_only=True
)

# Get categories (Level 0)
ws = wb['Categories']
categories = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[1] and row[2]:
        categories.append({'id': str(row[1]), 'name': str(row[2])})

print('=== TOP-LEVEL CATEGORIES ===')
for c in categories:
    print(c['id'] + ' - ' + c['name'])

# Get Level 1 processes from Combined sheet
ws2 = wb['Combined']
level1 = []
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0:
        continue
    hierarchy_id = str(row[1]) if row[1] else ''
    name = str(row[2]) if row[2] else ''
    parts = hierarchy_id.split('.')
    if len(parts) == 2 and parts[1] != '0':
        level1.append({'id': hierarchy_id, 'name': name, 'parent': parts[0] + '.0'})

print('\n=== LEVEL 1 PROCESSES (' + str(len(level1)) + ' total) ===')
for p in level1:
    print(p['id'] + ' - ' + p['name'] + ' (parent: ' + p['parent'] + ')')

# Output as JSON for data.js
output = {
    'categories': categories,
    'level1': level1
}
with open('data/apqc_capabilities.json', 'w') as f:
    json.dump(output, f, indent=2)
print('\nJSON written to data/apqc_capabilities.json')
